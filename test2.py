import pandas as pd

from openpyxl import load_workbook
import xlwt, xlrd
from xlutils.copy import copy

def write_excel(filename, datadict):
    # book = xlwt.Workbook()
    excel_name = filename + '.xlsx'
    for key in datadict.keys():
        sheet_name = key + '_sheet'
        # df = 'df_'+ key
        df = pd.DataFrame(data=datadict[key])

        with pd.ExcelWriter(excel_name) as writer:

            book = load_workbook(writer.path)
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name,header=0,index=None)

            writer.save()
            writer.close()


def write_excel_xls(path, data, sec):
    # index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    # sheet_name = []
    for key in data.keys():
        name = key         # sheet_name.append(name)
        sheet = workbook.add_sheet(name)  # 在工作簿中新建一个表格
        sheet.write(0,0,sec)
        for i in range(len(data[key])):
            sheet.write(0, i+1, data[key][i])
            
    # for i in range(0, index):
    #     for j in range(0, len(value[i])):
    #         sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）

    workbook.save(path)  # 保存工作簿
    print("xls格式表格写入数据成功！")

def write_excel_xls_append(path, data, sec):
    # index = len(value)  # 获取需要写入数据的行数
    
    # print(sheets)
    for key in data.keys():
        workbook = xlrd.open_workbook(path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        if key not in sheets:
            sheet = workbook.add_sheet(key)
            sheet.write(0, 0 ,sec)
            for i in range(len(data[key])):
                sheet(0, i + 1, data[key][i])
        else:
            worksheet = workbook.sheet_by_name(key)
            rows_old = worksheet.nrows
            new_workbook = copy(workbook)
            new_worksheet = new_workbook.get_sheet(key)
            new_worksheet.write(rows_old, 0, sec)
            for i in range(len(data[key])):
                new_worksheet.write(rows_old, i+1, data[key][i])
            new_workbook.save(path)
        print("xls格式表格【追加】写入数据成功！")



    

if __name__ == "__main__":
       data = {}
       data['READ'] = ['a','b','c']
       data['update'] = ['1','2','3']
       data['CLEAN'] = ['q', 'w', 'e']
       write_excel_xls('test1.xls', data, 10)
       data2 = {}
       data2['READ'] = ['1', '2', '3']
       data2['update'] = ['a', 'b', 'c']
       data2['CLEAN'] = ['q', 'w', 'e']
       write_excel_xls_append('test1.xls', data2, 10)
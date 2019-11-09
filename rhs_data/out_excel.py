import os, sys, re, codecs, xlwt
from pathlib import Path
import openpyxl
def data_dir():
    if (len(sys.argv) == 2):
        path = sys.argv[1]
        if os.path.isdir(path):
            print("Process file-path: %s" % path)
            print("Process file-name: All" )
            file_all = 1
        elif os.path.isfile(path):
            print("Process file-name: %s" % path )
            file_all = 0
        else:
            print("输入文件格式不正确")
    else:
        print("请输入文件参数")
        print(r"Example:python D:\脚本\deflation_shell\process_data\out_excel.py C:\Users\12093\Desktop\Deflation\colocat_deflation ")
    return path, file_all

def data_process(path):
    data = {}
    n=0
    tablename = ['TimeStap', 'Target address', 'RTT', 'initial address']
    f = codecs.open(path, 'rb', encoding='utf-8')
    for name in tablename:
        data[name] =  []
    line = f.readline()
    while line:
        n = n + 1
        data_p = line.split()
        # data_p = map(float, data_p)
        data_p = list(data_p)
        for i in range(len(tablename)):
            data[tablename[i]].append(data_p[i])
        line = f.readline()
    f.close()
    return data, n


def write_execl(path, tag):
    # excel表头
    tablename = ['TimeStap', 'Target address', 'RTT', 'initial address']
    if tag == 1:
        file_list = os.listdir(path)
        for name in file_list:
            file_path = path + '\\' + name
            data,n = data_process(file_path)
            book = openpyxl.Workbook()
            sheet = book.create_sheet(index=0)
            for i in range(len(data)):
                l = data.get(tablename[i])
                for j in range(n):
                    # sheet.write(j, i, l[j])
                    sheet.cell(j+1,i+1).value = l[j]
            l_name = name.split(".")
            book.save('%s.xlsx' % l_name[0])
    elif tag == 0:
        data,n = data_process(path)
        book = openpyxl.Workbook()
        sheet = book.create_sheet(index=0)
        for i in range(len(data)):
            l = data.get(tablename[i])
            for j in range(n):
                sheet.cell(j+1,i+1).value = l[j]
        l_name = path.split('\\')
        print(l_name)
        excel_name = l_name[-1].split('.')
        book.save('..\\%s.xlsx' % excel_name[0])

if __name__ == "__main__":
    path, tag= data_dir()
    write_execl(path, tag)
